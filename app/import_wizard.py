from collections import Counter

from openpyxl import Workbook, load_workbook

from .models import AcademicClass, AcademicYear, Exam, SchoolClass, Student, Subject


STUDENT_HEADERS = ["student_id", "full_name", "mother_name", "phone", "class", "academic_year"]
RESULT_HEADERS = ["student_id", "subject", "score"]


def student_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(STUDENT_HEADERS)
    ws.append(["3007", "Student Name", "Mother Name", "+252...", "Form 2", "2025-2026"])
    return wb


def result_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(RESULT_HEADERS)
    ws.append(["3007", "Math", 95])
    return wb


def preview_students(file):
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    headers = normalize_headers(ws[1])
    rows = []
    errors = []
    required = ["student_id", "full_name", "class", "academic_year"]
    missing = [header for header in required if header not in headers]
    if missing:
        errors.append(f"Missing required columns: {', '.join(missing)}")
        return rows, errors

    ids = []
    for index, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = row_dict(headers, values)
        if not data.get("student_id"):
            continue
        data["row"] = index
        data["student_id"] = str(data["student_id"]).strip()
        data["full_name"] = str(data.get("full_name") or "").strip()
        data["mother_name"] = str(data.get("mother_name") or "").strip()
        data["phone"] = str(data.get("phone") or "").strip()
        data["class"] = str(data.get("class") or "").strip()
        data["academic_year"] = str(data.get("academic_year") or "").strip()
        ids.append(data["student_id"])
        validate_student_row(data, errors)
        rows.append(data)

    for duplicate in duplicates(ids):
        errors.append(f"Duplicate Student ID in file: {duplicate}")
    return rows, errors


def preview_results(file, exam_id):
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    headers = normalize_headers(ws[1])
    rows = []
    errors = []
    missing = [header for header in RESULT_HEADERS if header not in headers]
    if missing:
        errors.append(f"Missing required columns: {', '.join(missing)}")
        return rows, errors
    exam = Exam.query.get(exam_id)
    if not exam:
        errors.append("Selected exam does not exist.")
        return rows, errors
    seen = []
    for index, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = row_dict(headers, values)
        if not data.get("student_id"):
            continue
        data["row"] = index
        data["student_id"] = str(data["student_id"]).strip()
        data["subject"] = str(data.get("subject") or "").strip()
        data["score"] = data.get("score")
        data["exam_id"] = int(exam_id)
        seen.append((data["student_id"], data["subject"]))
        validate_result_row(data, errors)
        rows.append(data)
    for student_id, subject in duplicates(seen):
        errors.append(f"Duplicate result in file: {student_id} / {subject}")
    return rows, errors


def normalize_headers(row):
    return [str(cell.value).strip().lower() if cell.value is not None else "" for cell in row]


def row_dict(headers, values):
    return {header: value for header, value in zip(headers, values) if header}


def duplicates(values):
    return [value for value, count in Counter(values).items() if count > 1]


def validate_student_row(data, errors):
    row = data["row"]
    if not data["full_name"]:
        errors.append(f"Row {row}: full_name is required.")
    if not SchoolClass.query.filter_by(name=data["class"]).first() and not AcademicClass.query.filter_by(name=data["class"]).first():
        errors.append(f"Row {row}: class does not exist: {data['class']}")
    if not AcademicYear.query.filter_by(name=data["academic_year"]).first():
        errors.append(f"Row {row}: academic year does not exist: {data['academic_year']}")


def validate_result_row(data, errors):
    row = data["row"]
    if not Student.query.filter_by(student_code=data["student_id"]).first():
        errors.append(f"Row {row}: student ID does not exist: {data['student_id']}")
    if not Subject.query.filter_by(name=data["subject"]).first():
        errors.append(f"Row {row}: subject does not exist: {data['subject']}")
    try:
        score = float(data["score"])
    except (TypeError, ValueError):
        errors.append(f"Row {row}: score must be a number.")
        return
    if score < 0 or score > 100:
        errors.append(f"Row {row}: score must be between 0 and 100.")
