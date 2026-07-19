from collections import defaultdict
from decimal import Decimal
import math
from tempfile import NamedTemporaryFile
from datetime import date

from flask import Blueprint, abort, current_app, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from sqlalchemy import or_ as db_or

from . import db
from .audit import audit
from .cloudinary_service import upload_image
from .import_wizard import preview_students, student_template
from .models import AcademicYear, AcademicClass, AcademicLevel, AcademicSection, Exam, GradeScale, IncidentReport, Result, SchoolClass, Setting, Student, Subject, LabelTranslation
from .permissions import can, enforce_endpoint_permission
from .security import ALLOWED_PHOTOS, ALLOWED_SHEETS, allowed_file
from .services import get_label, get_settings, grade_for, result_payload

advanced_results_bp = Blueprint("admin_advanced_results", __name__)


def ordinal(value):
    suffix = "th"
    if value % 100 not in (11, 12, 13):
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(value % 10, "th")
    return f"{value}{suffix}"


@advanced_results_bp.before_request
@login_required
def require_login():
    enforce_endpoint_permission()
    ensure_results_label_seeds()


@advanced_results_bp.app_context_processor
def inject_results_hub_helpers():
    return {"rh_label": get_label}


RESULTS_LABEL_SEEDS = [
    ("hub.brand.school", "so", "Dugsiga — Nidaamka Maamulka", "Shell"),
    ("hub.brand.subtitle", "so", "Results Hub", "Shell"),
    ("hub.user.role", "so", "Maamule · Super Admin", "Shell"),
    ("hub.tab.setup", "so", "Setup", "Shell"),
    ("hub.tab.dashboard", "so", "Dashboard", "Shell"),
    ("hub.tab.entry", "so", "Result Entry", "Shell"),
    ("hub.tab.roster", "so", "Liiska Fasalka", "Shell"),
    ("hub.tab.analytics", "so", "Analytics", "Shell"),
    ("hub.tab.grades", "so", "Grade Mgmt", "Shell"),
    ("hub.tab.settings", "so", "Settings", "Shell"),
    ("dashboard.eyebrow", "so", "Guudmarka", "Dashboard"),
    ("dashboard.title", "so", "Results Dashboard", "Dashboard"),
    ("dashboard.year", "so", "Sanad Dugsiyeedka", "Dashboard"),
    ("dashboard.exam_type", "so", "Exam Type", "Dashboard"),
    ("dashboard.total_students", "so", "Wadarta Ardayda", "Dashboard"),
    ("dashboard.subjects_entered", "so", "Maadooyinka la Xareeyay", "Dashboard"),
    ("dashboard.completion", "so", "Buuxinta Natiijooyinka", "Dashboard"),
    ("dashboard.active_classes", "so", "Fasallada Firfircoon", "Dashboard"),
    ("dashboard.step_level", "so", "Dooro Level", "Dashboard"),
    ("dashboard.step_class", "so", "Dooro Fasalka", "Dashboard"),
    ("dashboard.no_classes", "so", "Fasallo lama helin level-ka la doortay.", "Dashboard"),
    ("dashboard.select_exam_title", "so", "Dooro imtixaan si aad u aragto dashboard-ka.", "Dashboard"),
    ("entry.eyebrow", "so", "Form 4 · Bileedka 2aad · 2025-2026", "Result Entry"),
    ("entry.title", "so", "Whole-Class Result Entry", "Result Entry"),
    ("entry.subtitle", "so", "Geli dhibcaha ardayda oo dhan hal mar — validation-ka max_score ayaa toos u socda.", "Result Entry"),
    ("entry.unsaved", "so", "isbedel oo aan la kaydin", "Result Entry"),
    ("entry.draft", "so", "Draft", "Result Entry"),
    ("entry.published", "so", "Published", "Result Entry"),
    ("entry.save_all", "so", "Save All", "Result Entry"),
    ("entry.id", "so", "ID", "Result Entry"),
    ("entry.name", "so", "Magaca", "Result Entry"),
    ("entry.reset", "so", "Reset", "Result Entry"),
    ("entry.students", "so", "Students", "Result Entry"),
    ("entry.subjects", "so", "Subjects", "Result Entry"),
    ("roster.eyebrow", "so", "Form 4 · Bileedka 2aad · 2025-2026", "Roster"),
    ("roster.title", "so", "Liiska Natiijada Fasalka", "Roster"),
    ("roster.students", "so", "arday", "Roster"),
    ("roster.export_excel", "so", "Export Excel", "Roster"),
    ("roster.export_pdf", "so", "Export PDF", "Roster"),
    ("roster.search", "so", "Raadi arday ID ama magac...", "Roster"),
    ("roster.student_name", "so", "Magaca", "Roster"),
    ("roster.total", "so", "Total", "Roster"),
    ("roster.grade", "so", "Grade", "Roster"),
    ("analytics.eyebrow", "so", "Form 4 · Bileedka 2aad · 2025-2026", "Analytics"),
    ("analytics.title", "so", "Analytics — Infographic", "Analytics"),
    ("analytics.average", "so", "Celceliska Fasalka", "Analytics"),
    ("analytics.gpa", "so", "GPA Fasalka", "Analytics"),
    ("analytics.pass_rate", "so", "Pass Rate", "Analytics"),
    ("analytics.grade_distribution", "so", "Grade Distribution", "Analytics"),
    ("analytics.subject_performance", "so", "Subject-wise Performance", "Analytics"),
    ("analytics.exam_trend", "so", "Exam-Type Trend", "Analytics"),
    ("analytics.performers", "so", "Top & Bottom Performers", "Analytics"),
    ("analytics.top", "so", "Top 5", "Analytics"),
    ("analytics.bottom", "so", "Bottom 5", "Analytics"),
    ("grades.eyebrow", "so", "Grade Management", "Grade Management"),
    ("grades.title", "so", "Grade Management — Simplified", "Grade Management"),
    ("grades.subtitle", "so", "Academic Year → Exam Type → hal scale oo si toos ah loogu dabaqo dhammaan fasallada imtixaankan qaaday.", "Grade Management"),
    ("grades.applies_to", "so", "Applies To", "Grade Management"),
    ("grades.all_classes", "so", "All Classes", "Grade Management"),
    ("grades.generate", "so", "Generate Scale", "Grade Management"),
    ("grades.grade", "so", "Grade", "Grade Management"),
    ("grades.min", "so", "Min", "Grade Management"),
    ("grades.max", "so", "Max", "Grade Management"),
    ("grades.point", "so", "Point", "Grade Management"),
    ("grades.comment", "so", "Comment", "Grade Management"),
    ("grades.preview", "so", "Preview", "Grade Management"),
    ("grades.save", "so", "Save Grade Scales", "Grade Management"),
    ("settings.eyebrow", "so", "Full Customization", "Settings"),
    ("settings.title", "so", "Results Settings", "Settings"),
    ("settings.subtitle", "so", "Halkan waxaad ka bedeli kartaa habka module-ka Results u shaqeeyo — gudaha Results-ka, ma aha Settings-ka guud.", "Settings"),
    ("settings.labels_title", "so", "Label & Language Customization", "Settings"),
    ("settings.labels_desc", "so", "Halkan waxaad ka bedeli kartaa ereyga/label kasta oo systemka ka muuqda.", "Settings"),
    ("settings.default_language", "so", "Luuqadda Default-ka ah", "Settings"),
    ("settings.add_language", "so", "Ku dar Luuqad", "Settings"),
    ("settings.label_key", "so", "Label Key", "Settings"),
    ("settings.context", "so", "Meesha ka muuqato", "Settings"),
    ("settings.save_labels", "so", "Save Labels", "Settings"),
]


def ensure_results_label_seeds():
    if current_app.config.get("_results_label_seeds_checked"):
        return
    changed = False
    for label_key, language_code, text_value, context in RESULTS_LABEL_SEEDS:
        exists = LabelTranslation.query.filter_by(label_key=label_key, language_code=language_code).first()
        if exists:
            continue
        db.session.add(LabelTranslation(
            label_key=label_key,
            language_code=language_code,
            text_value=text_value,
            context=context,
        ))
        changed = True
    if changed:
        db.session.commit()
    current_app.config["_results_label_seeds_checked"] = True


def get_default_academic_year(year_id=None):
    """Return the requested year, otherwise the latest active academic year."""
    if year_id:
        return db.session.get(AcademicYear, year_id)
    return (
        AcademicYear.query.filter_by(is_current=True)
        .order_by(AcademicYear.name.desc(), AcademicYear.id.desc())
        .first()
        or AcademicYear.query.order_by(AcademicYear.name.desc(), AcademicYear.id.desc()).first()
    )


def get_latest_exam_for_year(academic_year):
    """Return the latest active exam for the selected academic year, falling back to any latest exam."""
    if not academic_year:
        return None
    return (
        Exam.query.filter_by(academic_year_id=academic_year.id, is_active=True)
        .order_by(Exam.id.desc())
        .first()
        or
        Exam.query.filter_by(academic_year_id=academic_year.id)
        .order_by(Exam.id.desc())
        .first()
    )


def subjects_for_scope(exam, level_id=None, class_id=None):
    """Return subjects for the effective academic level of the current scope."""
    effective_level_id = exam.academic_level_id if exam else None
    if not effective_level_id:
        effective_level_id = level_id
    if not effective_level_id and class_id:
        academic_class = db.session.get(AcademicClass, class_id)
        effective_level_id = academic_class.academic_level_id if academic_class else None

    query = Subject.query
    if effective_level_id:
        scoped_subjects = query.filter_by(academic_level_id=effective_level_id).all()
        if scoped_subjects:
            return sorted(scoped_subjects, key=lambda subject: (subject.sort_order, subject.name))
        query = Subject.query.filter(Subject.academic_level_id.is_(None))
    return sorted(query.all(), key=lambda subject: (subject.sort_order, subject.name))


def students_for_scope_query(academic_year_id, level_id=None, class_id=None, section_id=None, exam=None):
    """Return students using the unified Results Hub academic hierarchy, with legacy compatibility."""
    query = Student.query.filter_by(academic_year_id=academic_year_id)

    effective_level_id = level_id or (exam.academic_level_id if exam else None)
    effective_class_id = class_id or (exam.academic_class_id if exam else None)
    effective_section_id = section_id or (exam.academic_section_id if exam else None)

    section = db.session.get(AcademicSection, effective_section_id) if effective_section_id else None
    if section and not effective_class_id:
        effective_class_id = section.academic_class_id

    academic_class = db.session.get(AcademicClass, effective_class_id) if effective_class_id else None
    if academic_class and not effective_level_id:
        effective_level_id = academic_class.academic_level_id

    if effective_class_id:
        class_filters = [Student.academic_class_id == effective_class_id]
        if academic_class:
            legacy_class = SchoolClass.query.filter_by(name=academic_class.name).first()
            if legacy_class:
                class_filters.append(Student.class_id == legacy_class.id)
        query = query.filter(db_or(*class_filters))
    elif effective_level_id:
        level_filters = [Student.academic_level_id == effective_level_id]
        academic_level = db.session.get(AcademicLevel, effective_level_id)
        if academic_level:
            level_filters.append(Student.level == academic_level.name)
        query = query.filter(db_or(*level_filters))

    if effective_section_id:
        section_filters = [Student.academic_section_id == effective_section_id]
        if section:
            section_filters.append(Student.section == section.name)
        query = query.filter(db_or(*section_filters))

    return query


def rank_student_in_scope(student, academic_year_id, exam, subjects, level_id=None, class_id=None, section_id=None):
    """Return the student's rank within the selected academic scope."""
    if not student or not exam or not subjects:
        return "-"

    scoped_students = students_for_scope_query(
        academic_year_id,
        level_id=level_id,
        class_id=class_id,
        section_id=section_id,
        exam=exam,
    ).all()
    if not scoped_students:
        return "-"

    student_ids = [row.id for row in scoped_students]
    results = Result.query.filter(
        Result.exam_id == exam.id,
        Result.student_id.in_(student_ids),
    ).all()
    results_by_student = {}
    for result in results:
        results_by_student.setdefault(result.student_id, {})[result.subject_id] = result

    ranked = []
    for scoped_student in scoped_students:
        total_score = 0
        total_max = 0
        student_results = results_by_student.get(scoped_student.id, {})
        for subject in subjects:
            result = student_results.get(subject.id)
            total_score += float(result.score) if result else 0
            total_max += float(subject.max_score)
        percentage = (total_score / total_max * 100) if total_max > 0 else 0
        ranked.append((scoped_student.id, percentage, total_score))

    ranked.sort(key=lambda item: (item[1], item[2]), reverse=True)
    for index, (student_id, _percentage, _total_score) in enumerate(ranked, start=1):
        if student_id == student.id:
            return index
    return "-"


@advanced_results_bp.route("/")
def dashboard():
    """Results Hub landing page - kept as a compatibility route."""
    return redirect(url_for("admin_advanced_results.new_dashboard"))


@advanced_results_bp.route("/new-setup")
def new_setup():
    """Setup wizard - Master Configuration for the entire Results system"""
    level_id = int_or_none(request.args.get("level_id"))
    
    # Get selected year (current year by default)
    selected_year = AcademicYear.query.filter_by(is_current=True).first()
    selected_level = db.session.get(AcademicLevel, level_id) if level_id else None
    
    # Get all data for Setup wizard
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    exams = Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).all() if selected_year else []
    levels = AcademicLevel.query.filter_by(is_active=True).order_by(AcademicLevel.sort_order).all()
    subjects = Subject.query.order_by(Subject.sort_order).all()
    classes = AcademicClass.query.order_by(AcademicClass.name).all()
    
    # Calculate step completion states
    step_states = {
        'academic_year': len(years) > 0,
        'exam_type': len(exams) > 0,
        'levels_classes': False,  # Will check below
        'subjects': False  # Will check below
    }
    
    # Check if any level has at least one class
    for level in levels:
        if level.classes.count() > 0:
            step_states['levels_classes'] = True
            break
    
    # Check if subjects are configured
    if subjects:
        step_states['subjects'] = True
    
    # Determine current active step
    if not step_states['academic_year']:
        current_step = 'academic_year'
    elif not step_states['exam_type']:
        current_step = 'exam_type'
    elif not step_states['levels_classes']:
        current_step = 'levels_classes'
    elif not step_states['subjects']:
        current_step = 'subjects'
    else:
        current_step = 'subjects'  # All complete
    
    return render_template(
        "admin/results_setup.html",
        years=years,
        exams=exams,
        levels=levels,
        selected_year=selected_year,
        selected_level=selected_level,
        subjects=subjects,
        classes=classes,
        settings=get_settings(),
        step_states=step_states,
        current_step=current_step,
    )


# Level CRUD Routes for Setup Page
@advanced_results_bp.route("/setup/levels/add", methods=["POST"])
def setup_add_level():
    """Add a new academic level from Setup page"""
    name = request.form.get("name", "").strip()
    
    if not name:
        flash("Level name is required.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    # Get max sort_order
    from sqlalchemy import func
    max_order = db.session.query(func.max(AcademicLevel.sort_order)).scalar() or 0
    
    level = AcademicLevel(
        name=name,
        sort_order=max_order + 1,
        is_active=True
    )
    
    db.session.add(level)
    audit("System Setup", f"Added academic level: {level.name}")
    db.session.commit()
    flash("Level added successfully.", "success")
    return redirect(url_for("admin_advanced_results.new_setup"))


@advanced_results_bp.route("/setup/levels/<int:level_id>/edit", methods=["POST"])
def setup_edit_level(level_id):
    """Edit an academic level from Setup page"""
    level = db.session.get(AcademicLevel, level_id)
    if not level:
        flash("Level not found.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    name = request.form.get("name", "").strip()
    if not name:
        flash("Level name is required.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    level.name = name
    audit("System Setup", f"Edited academic level: {level.name}")
    db.session.commit()
    flash("Level updated successfully.", "success")
    return redirect(url_for("admin_advanced_results.new_setup"))


@advanced_results_bp.route("/setup/levels/<int:level_id>/delete", methods=["POST"])
def setup_delete_level(level_id):
    """Delete an academic level from Setup page"""
    level = db.session.get(AcademicLevel, level_id)
    if not level:
        flash("Level not found.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    # Check if level has classes
    if level.classes.count() > 0:
        flash("Cannot delete level with existing classes.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    db.session.delete(level)
    audit("System Setup", f"Deleted academic level: {level.name}")
    db.session.commit()
    flash("Level deleted successfully.", "success")
    return redirect(url_for("admin_advanced_results.new_setup"))


# Class CRUD Routes for Setup Page
@advanced_results_bp.route("/setup/classes/add", methods=["POST"])
def setup_add_class():
    """Add a new academic class from Setup page"""
    academic_level_id = int(request.form.get("academic_level_id"))
    name = request.form.get("name", "").strip()
    
    if not name:
        flash("Class name is required.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    level = db.session.get(AcademicLevel, academic_level_id)
    if not level:
        flash("Academic level not found.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    # Get max sort_order for this level
    from sqlalchemy import func
    max_order = db.session.query(func.max(AcademicClass.sort_order)).filter_by(
        academic_level_id=academic_level_id
    ).scalar() or 0
    
    cls = AcademicClass(
        academic_level_id=academic_level_id,
        name=name,
        sort_order=max_order + 1,
        is_active=True
    )
    
    db.session.add(cls)
    audit("System Setup", f"Added academic class: {cls.name}")
    db.session.commit()
    flash("Class added successfully.", "success")
    return redirect(url_for("admin_advanced_results.new_setup"))


@advanced_results_bp.route("/setup/classes/<int:class_id>/edit", methods=["POST"])
def setup_edit_class(class_id):
    """Edit an academic class from Setup page"""
    cls = db.session.get(AcademicClass, class_id)
    if not cls:
        flash("Class not found.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    name = request.form.get("name", "").strip()
    if not name:
        flash("Class name is required.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    cls.name = name
    audit("System Setup", f"Edited academic class: {cls.name}")
    db.session.commit()
    flash("Class updated successfully.", "success")
    return redirect(url_for("admin_advanced_results.new_setup"))


@advanced_results_bp.route("/setup/classes/<int:class_id>/delete", methods=["POST"])
def setup_delete_class(class_id):
    """Delete an academic class from Setup page"""
    cls = db.session.get(AcademicClass, class_id)
    if not cls:
        flash("Class not found.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    # Check if class has sections
    if cls.sections.count() > 0:
        flash("Cannot delete class with existing sections.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    db.session.delete(cls)
    audit("System Setup", f"Deleted academic class: {cls.name}")
    db.session.commit()
    flash("Class deleted successfully.", "success")
    return redirect(url_for("admin_advanced_results.new_setup"))


@advanced_results_bp.route("/new-dashboard")
def new_dashboard():
    """New results dashboard with auto-selection of active academic year"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    level_id = int_or_none(request.args.get("level_id"))
    
    selected_year = get_default_academic_year(year_id)
    selected_exam = db.session.get(Exam, exam_id) if exam_id else get_latest_exam_for_year(selected_year)
    if selected_exam and selected_year and selected_exam.academic_year_id != selected_year.id:
        selected_exam = get_latest_exam_for_year(selected_year)
    selected_level = db.session.get(AcademicLevel, level_id) if level_id else None
    
    # Get all years and exams for selectors
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    exams = Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).all() if selected_year else []
    
    # Get all levels for level selector
    levels = AcademicLevel.query.filter_by(is_active=True).order_by(AcademicLevel.sort_order).all()
    
    stats = {}
    class_cards = []
    
    if selected_exam:
        stats = build_dashboard_stats(selected_exam)
        class_cards = build_class_cards(selected_exam, level_filter=selected_level)
    
    return render_template(
        "admin/results_dashboard.html",
        years=years,
        exams=exams,
        levels=levels,
        selected_year=selected_year,
        selected_exam=selected_exam,
        selected_level=selected_level,
        stats=stats,
        class_cards=class_cards,
        settings=get_settings(),
    )


@advanced_results_bp.route("/class-roster")
def class_roster():
    """Class roster view with all students and their results"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    level_id = int_or_none(request.args.get("level_id"))
    class_id = int_or_none(request.args.get("class_id"))
    section_id = int_or_none(request.args.get("section_id"))
    student_id = int_or_none(request.args.get("student_id"))
    search_query = request.args.get("q", "").strip()
    
    selected_year = get_default_academic_year(year_id)
    selected_exam = db.session.get(Exam, exam_id) if exam_id else get_latest_exam_for_year(selected_year)
    if selected_exam and selected_year and selected_exam.academic_year_id != selected_year.id:
        selected_exam = get_latest_exam_for_year(selected_year)
    
    if not selected_year:
        flash("Please select an academic year.", "warning")
        return redirect(url_for("admin_advanced_results.new_dashboard"))

    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    exams = Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).all() if selected_year else []
    levels = AcademicLevel.query.filter_by(is_active=True).order_by(AcademicLevel.sort_order).all()
    level_id = level_id or (selected_exam.academic_level_id if selected_exam else None)
    class_id = class_id or (selected_exam.academic_class_id if selected_exam else None)
    section_id = section_id or (selected_exam.academic_section_id if selected_exam else None)
    classes = AcademicClass.query.filter_by(academic_level_id=level_id, is_active=True).order_by(AcademicClass.sort_order, AcademicClass.name).all() if level_id else []
    sections = AcademicSection.query.filter_by(academic_class_id=class_id, is_active=True).order_by(AcademicSection.sort_order, AcademicSection.name).all() if class_id else []

    scope_info = {
        "level": db.session.get(AcademicLevel, level_id) if level_id else None,
        "class": db.session.get(AcademicClass, class_id) if class_id else None,
        "section": db.session.get(AcademicSection, section_id) if section_id else None,
    }
    
    # If no exam selected, show exam selection interface
    if not selected_exam:
        return render_template(
            "admin/class_roster.html",
            selected_year=selected_year,
            selected_exam=None,
            scope_info=scope_info,
            students=[],
            subjects=[],
            years=years,
            exams=exams,
            levels=levels,
            classes=classes,
            sections=sections,
            search_query=search_query,
            settings=get_settings(),
        )

    if not (level_id and class_id):
        return render_template(
            "admin/class_roster.html",
            selected_year=selected_year,
            selected_exam=selected_exam,
            scope_info=scope_info,
            students=[],
            subjects=[],
            years=years,
            exams=exams,
            levels=levels,
            classes=classes,
            sections=sections,
            search_query=search_query,
            settings=get_settings(),
        )
    
    student_query = students_for_scope_query(
        selected_year.id,
        level_id=level_id,
        class_id=class_id,
        section_id=section_id,
    )
    
    # Apply search filter
    if search_query:
        search_pattern = f"%{search_query}%"
        student_query = student_query.filter(
            db_or(
                Student.student_code.like(search_pattern),
                Student.full_name.like(search_pattern)
            )
        )
    
    students = student_query.order_by(Student.full_name).all()
    
    subjects = subjects_for_scope(selected_exam, level_id=level_id, class_id=class_id)
    
    # Build results data for each student
    roster_data = []
    for student in students:
        # Get results for this student and exam (only published results)
        results = Result.query.filter_by(student_id=student.id, exam_id=exam_id, is_published=True).all()
        results_dict = {r.subject_id: r for r in results}
        
        # Calculate totals and grades
        total_score = 0
        total_max = 0
        subject_data = []
        
        for subject in subjects:
            result = results_dict.get(subject.id)
            score = float(result.score) if result else 0
            max_score = float(subject.max_score)
            percentage = (score / max_score * 100) if max_score > 0 else 0
            
            total_score += score
            total_max += max_score
            
            # Get grade using exam-specific grade scale
            grade_info = grade_for(percentage, exam_id=exam_id)
            
            # Apply grade_override if present
            if result and result.grade_override:
                grade_info = dict(grade_info)
                grade_info["grade"] = result.grade_override
            
            subject_data.append({
                "subject": subject,
                "result": result,
                "score": score,
                "max_score": max_score,
                "percentage": percentage,
                "grade": grade_info,
            })
        
        overall_percentage = (total_score / total_max * 100) if total_max > 0 else 0
        overall_grade = grade_for(overall_percentage, exam_id=exam_id)
        
        # Calculate GP (grade point average)
        total_points = sum(s["grade"]["grade_point"] for s in subject_data if s["grade"]["grade_point"])
        gp = round(total_points / len(subject_data), 2) if subject_data else 0
        
        roster_data.append({
            "student": student,
            "subject_data": subject_data,
            "total_score": total_score,
            "total_max": total_max,
            "percentage": overall_percentage,
            "grade": overall_grade,
            "gp": gp,
        })
    
    return render_template(
        "admin/class_roster.html",
        selected_year=selected_year,
        selected_exam=selected_exam,
        scope_info=scope_info,
        students=roster_data,
        subjects=subjects,
        years=years,
        exams=exams,
        levels=levels,
        classes=classes,
        sections=sections,
        search_query=search_query,
        settings=get_settings(),
    )


@advanced_results_bp.route("/student-view")
def student_view():
    """Single student view with detailed results"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    student_id = int_or_none(request.args.get("student_id"))
    
    # Get selected year, exam, and student
    selected_year = db.session.get(AcademicYear, year_id)
    selected_exam = db.session.get(Exam, exam_id)
    student = db.session.get(Student, student_id)
    
    if not selected_year or not student:
        flash("Invalid selection.", "warning")
        return redirect(url_for("admin_advanced_results.new_dashboard"))
    if not selected_exam or selected_exam.academic_year_id != selected_year.id:
        selected_exam = get_latest_exam_for_year(selected_year)
    if not selected_exam:
        flash("Please select an exam type.", "warning")
        return redirect(url_for("admin_advanced_results.result_entry", year_id=selected_year.id))
    
    subjects = subjects_for_scope(
        selected_exam,
        level_id=student.academic_level_id,
        class_id=student.academic_class_id,
    )
    
    # Get results for this student and exam
    results = Result.query.filter_by(student_id=student.id, exam_id=selected_exam.id).all()
    results_dict = {r.subject_id: r for r in results}
    
    # Build subject data
    subject_data = []
    total_score = 0
    total_max = 0
    grade_distribution = {}
    
    for subject in subjects:
        result = results_dict.get(subject.id)
        score = float(result.score) if result else 0
        max_score = float(subject.max_score)
        percentage = (score / max_score * 100) if max_score > 0 else 0
        
        total_score += score
        total_max += max_score
        
        # Get grade using exam-specific grade scale
        grade_info = grade_for(percentage, exam_id=selected_exam.id)
        
        # Apply grade_override if present
        if result and result.grade_override:
            grade_info = dict(grade_info)
            grade_info["grade"] = result.grade_override
        grade_distribution[grade_info["grade"]] = grade_distribution.get(grade_info["grade"], 0) + 1
        
        subject_data.append({
            "subject": subject,
            "result": result,
            "score": score,
            "max_score": max_score,
            "pass_mark": round(max_score * 0.5, 2),
            "percentage": percentage,
            "grade": grade_info,
            "remark": grade_info.get("comment") or ("Pass" if grade_info.get("is_pass") else "Needs Improvement"),
        })
    
    overall_percentage = (total_score / total_max * 100) if total_max > 0 else 0
    overall_grade = grade_for(overall_percentage, exam_id=selected_exam.id)
    
    # Calculate GP
    total_points = sum(s["grade"]["grade_point"] for s in subject_data if s["grade"]["grade_point"])
    gp = round(total_points / len(subject_data), 2) if subject_data else 0
    status = "Passed" if overall_grade.get("is_pass") else "Failed"
    rank = rank_student_in_scope(
        student,
        selected_year.id,
        selected_exam,
        subjects,
        level_id=student.academic_level_id,
        class_id=student.academic_class_id,
        section_id=student.academic_section_id,
    )
    
    return render_template(
        "admin/student_view.html",
        years=AcademicYear.query.order_by(AcademicYear.name.desc()).all(),
        exams=Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).all(),
        selected_year=selected_year,
        selected_exam=selected_exam,
        student=student,
        subject_data=subject_data,
        total_score=total_score,
        total_max=total_max,
        percentage=overall_percentage,
        grade=overall_grade,
        gp=gp,
        rank=rank,
        status=status,
        grade_distribution=grade_distribution,
        settings=get_settings(),
    )


@advanced_results_bp.route("/bulk", methods=["POST"])
def bulk():
    ids = [int(value) for value in request.form.getlist("result_ids") if value.isdigit()]
    action = request.form.get("action", "")
    if not ids:
        flash("Select result rows first.", "warning")
        return redirect(url_for("admin_advanced_results.dashboard"))
    rows = Result.query.filter(Result.id.in_(ids)).all()
    if action == "publish":
        for row in rows:
            row.is_published = True
        audit("Result Publishing", f"Bulk published {len(rows)} result rows")
    elif action == "unpublish":
        for row in rows:
            row.is_published = False
        audit("Result Publishing", f"Bulk unpublished {len(rows)} result rows")
    elif action == "lock":
        for row in rows:
            row.student.is_result_locked = True
            row.student.lock_reason = "Locked from advanced results."
        audit("Result Locking", f"Bulk locked {len(rows)} result rows")
    elif action == "unlock":
        for row in rows:
            row.student.is_result_locked = False
            row.student.lock_reason = ""
        audit("Result Locking", f"Bulk unlocked {len(rows)} result rows")
    elif action == "delete":
        for row in rows:
            db.session.delete(row)
        audit("Result Publishing", f"Bulk deleted {len(rows)} result rows")
    else:
        abort(400)
    db.session.commit()
    flash("Bulk result action completed.", "success")
    return redirect(url_for("admin_advanced_results.dashboard"))


@advanced_results_bp.route("/export.xlsx")
def export_excel():
    """Original advanced results Excel export - kept for backward compatibility"""
    filters = result_filters()
    wb = Workbook()
    ws = wb.active
    ws.title = "Advanced Results"
    ws.append(["Academic Year", "Exam", "Level", "Class", "Section", "Student ID", "Student Name", "Subject", "Score", "Published", "Locked"])
    for row in result_query(filters).order_by(Result.updated_at.desc()).all():
        ws.append([
            row.exam.academic_year.name,
            row.exam.name,
            row.student.level or "",
            row.student.school_class.name,
            row.student.section or "",
            row.student.student_code,
            row.student.full_name,
            row.subject.name,
            float(row.score),
            row.is_published,
            row.student.is_result_locked,
        ])
    audit("Result Export", "Exported advanced results")
    db.session.commit()
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    return send_file(tmp.name, as_attachment=True, download_name="advanced_results.xlsx")


@advanced_results_bp.route("/export-student-pdf")
def export_student_pdf():
    """Export single student results as PDF"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    student_id = int_or_none(request.args.get("student_id"))
    
    selected_year = db.session.get(AcademicYear, year_id)
    selected_exam = db.session.get(Exam, exam_id)
    student = db.session.get(Student, student_id)
    
    if not selected_year or not selected_exam or not student:
        abort(404)
    
    # Get subjects and results
    subjects = Subject.query.filter_by(academic_level_id=selected_exam.academic_level_id).all() if selected_exam.academic_level_id else Subject.query.all()
    subjects = sorted(subjects, key=lambda s: (s.sort_order, s.name))
    
    results = Result.query.filter_by(student_id=student.id, exam_id=exam_id, is_published=True).all()
    results_dict = {r.subject_id: r for r in results}
    
    # Build data
    subject_data = []
    total_score = 0
    total_max = 0
    
    for subject in subjects:
        result = results_dict.get(subject.id)
        score = float(result.score) if result else 0
        max_score = float(subject.max_score)
        percentage = (score / max_score * 100) if max_score > 0 else 0
        
        total_score += score
        total_max += max_score
        
        grade_info = grade_for(percentage, exam_id=exam_id)
        
        # Apply grade_override if present
        if result and result.grade_override:
            grade_info = dict(grade_info)
            grade_info["grade"] = result.grade_override
        
        subject_data.append({
            "subject": subject,
            "score": score,
            "max_score": max_score,
            "percentage": percentage,
            "grade": grade_info,
        })
    
    overall_percentage = (total_score / total_max * 100) if total_max > 0 else 0
    overall_grade = grade_for(overall_percentage, exam_id=exam_id)
    
    total_points = sum(s["grade"]["grade_point"] for s in subject_data if s["grade"]["grade_point"])
    gp = round(total_points / len(subject_data), 2) if subject_data else 0
    
    settings = get_settings()
    
    return render_template(
        "admin/pdf/student_result_pdf.html",
        selected_year=selected_year,
        selected_exam=selected_exam,
        student=student,
        subject_data=subject_data,
        total_score=total_score,
        total_max=total_max,
        percentage=overall_percentage,
        grade=overall_grade,
        gp=gp,
        settings=settings,
        date=date.today(),
    )


@advanced_results_bp.route("/export-class-pdf")
def export_class_pdf():
    """Export class results as PDF mark sheet"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    level_id = int_or_none(request.args.get("level_id"))
    class_id = int_or_none(request.args.get("class_id"))
    section_id = int_or_none(request.args.get("section_id"))
    
    selected_year = db.session.get(AcademicYear, year_id)
    selected_exam = db.session.get(Exam, exam_id)
    
    if not selected_year or not selected_exam:
        abort(404)
    if not (level_id and class_id):
        abort(400)
    
    students = (
        students_for_scope_query(
            year_id,
            level_id=level_id,
            class_id=class_id,
            section_id=section_id,
        )
        .order_by(Student.full_name)
        .all()
    )
    subjects = subjects_for_scope(selected_exam, level_id=level_id, class_id=class_id)
    
    # Build roster data
    roster_data = []
    for student in students:
        results = Result.query.filter_by(student_id=student.id, exam_id=exam_id, is_published=True).all()
        results_dict = {r.subject_id: r for r in results}
        
        total_score = 0
        total_max = 0
        subject_data = []
        
        for subject in subjects:
            result = results_dict.get(subject.id)
            score = float(result.score) if result else 0
            max_score = float(subject.max_score)
            percentage = (score / max_score * 100) if max_score > 0 else 0
            
            total_score += score
            total_max += max_score
            
            grade_info = grade_for(percentage, exam_id=exam_id)
            
            # Apply grade_override if present
            if result and result.grade_override:
                grade_info = dict(grade_info)
                grade_info["grade"] = result.grade_override
            
            subject_data.append({
                "score": score,
                "percentage": percentage,
                "grade": grade_info,
            })
        
        overall_percentage = (total_score / total_max * 100) if total_max > 0 else 0
        overall_grade = grade_for(overall_percentage, exam_id=exam_id)
        
        roster_data.append({
            "student": student,
            "subject_data": subject_data,
            "total_score": total_score,
            "total_max": total_max,
            "percentage": overall_percentage,
            "grade": overall_grade,
        })

    ranked_data = sorted(roster_data, key=lambda row: (row["percentage"], row["total_score"]), reverse=True)
    rank_lookup = {row["student"].id: index for index, row in enumerate(ranked_data, start=1)}
    for row in roster_data:
        row["rank"] = rank_lookup.get(row["student"].id, 0)
        row["rank_label"] = ordinal(row["rank"]) if row["rank"] else "-"
    
    # Calculate class stats
    class_average = round(sum(r["percentage"] for r in roster_data) / len(roster_data), 2) if roster_data else 0
    highest_total = round(max((row["total_score"] for row in roster_data), default=0), 2)
    lowest_total = round(min((row["total_score"] for row in roster_data), default=0), 2)
    average_total = round(sum(row["total_score"] for row in roster_data) / len(roster_data), 2) if roster_data else 0
    passed_count = sum(1 for row in roster_data if row["grade"].get("is_pass"))
    failed_count = len(roster_data) - passed_count
    pass_rate = round((passed_count / len(roster_data) * 100), 2) if roster_data else 0
    highest_score = max((row["percentage"] for row in roster_data), default=0)
    lowest_score = min((row["percentage"] for row in roster_data), default=0)
    subject_stats = []
    for index, subject in enumerate(subjects):
        scores = [float(row["subject_data"][index]["score"]) for row in roster_data if index < len(row["subject_data"])]
        subject_stats.append({
            "subject": subject,
            "highest": round(max(scores), 2) if scores else 0,
            "lowest": round(min(scores), 2) if scores else 0,
            "average": round(sum(scores) / len(scores), 2) if scores else 0,
        })
    students_per_page = 15
    total_pages = max(1, math.ceil(len(roster_data) / students_per_page))
    
    # Get scope info
    scope_info = {}
    if level_id:
        scope_info["level"] = db.session.get(AcademicLevel, level_id)
    if class_id:
        scope_info["class"] = db.session.get(AcademicClass, class_id)
    if section_id:
        scope_info["section"] = db.session.get(AcademicSection, section_id)
    
    settings = get_settings()
    
    return render_template(
        "admin/pdf/class_mark_sheet_pdf.html",
        selected_year=selected_year,
        selected_exam=selected_exam,
        scope_info=scope_info,
        students=roster_data,
        subjects=subjects,
        subject_stats=subject_stats,
        class_average=class_average,
        highest_total=highest_total,
        lowest_total=lowest_total,
        average_total=average_total,
        passed_count=passed_count,
        failed_count=failed_count,
        pass_rate=pass_rate,
        highest_score=round(highest_score, 2),
        lowest_score=round(lowest_score, 2),
        completed_count=len(roster_data),
        students_per_page=students_per_page,
        total_pages=total_pages,
        settings=settings,
        generated_by=current_user.full_name or current_user.username,
        date=date.today(),
    )


@advanced_results_bp.route("/export-class-excel")
def export_class_excel():
    """Export class results as Excel with conditional formatting"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    level_id = int_or_none(request.args.get("level_id"))
    class_id = int_or_none(request.args.get("class_id"))
    section_id = int_or_none(request.args.get("section_id"))
    
    selected_year = db.session.get(AcademicYear, year_id)
    selected_exam = db.session.get(Exam, exam_id)
    
    if not selected_year or not selected_exam:
        abort(404)
    if not (level_id and class_id):
        abort(400)
    
    students = (
        students_for_scope_query(
            year_id,
            level_id=level_id,
            class_id=class_id,
            section_id=section_id,
        )
        .order_by(Student.full_name)
        .all()
    )
    subjects = subjects_for_scope(selected_exam, level_id=level_id, class_id=class_id)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Class Results"
    
    # Header row
    headers = ["ID", "Student Name", "Mother's Name", "Class"]
    for subject in subjects:
        headers.append(subject.name)
    headers.extend(["Total", "%", "Grade", "GP"])
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    row_num = 2
    for student in students:
        results = Result.query.filter_by(student_id=student.id, exam_id=exam_id, is_published=True).all()
        results_dict = {r.subject_id: r for r in results}
        
        total_score = 0
        total_max = 0
        row_data = [
            student.student_code,
            student.full_name,
            student.mother_name or "",
            f"{student.academic_class.name if student.academic_class else student.level or '-'}{' - ' + student.academic_section.name if student.academic_section else ''}"
        ]
        
        for subject in subjects:
            result = results_dict.get(subject.id)
            score = float(result.score) if result else 0
            max_score = float(subject.max_score)
            percentage = (score / max_score * 100) if max_score > 0 else 0
            
            total_score += score
            total_max += max_score
            
            row_data.append(score)
        
        overall_percentage = (total_score / total_max * 100) if total_max > 0 else 0
        overall_grade = grade_for(overall_percentage, exam_id=exam_id)
        
        total_points = 0
        for subject in subjects:
            result = results_dict.get(subject.id)
            if result:
                percentage = (float(result.score) / float(subject.max_score) * 100) if subject.max_score else 0
                grade_info = grade_for(percentage, exam_id=exam_id)
                # Apply grade_override if present
                if result.grade_override:
                    grade_info = dict(grade_info)
                    grade_info["grade"] = result.grade_override
                total_points += grade_info["grade_point"]
        
        gp = round(total_points / len(subjects), 2) if subjects else 0
        
        row_data.extend([total_score, overall_percentage, overall_grade["grade"], gp])
        ws.append(row_data)
        
        # Apply conditional formatting to percentage column
        percentage_col = len(subjects) + 5  # ID, Name, Mother, Class + subjects + Total
        percentage_cell = ws.cell(row=row_num, column=percentage_col)
        
        if overall_percentage >= 70:
            percentage_cell.fill = PatternFill(start_color="E7F1EA", end_color="E7F1EA", fill_type="solid")
        elif overall_percentage >= 50:
            percentage_cell.fill = PatternFill(start_color="FBF3DE", end_color="FBF3DE", fill_type="solid")
        else:
            percentage_cell.fill = PatternFill(start_color="F7E9E7", end_color="F7E9E7", fill_type="solid")
        
        row_num += 1
    
    audit("Result Export", f"Exported class results for exam {exam_id}")
    db.session.commit()
    
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    return send_file(tmp.name, as_attachment=True, download_name="class_results.xlsx")


@advanced_results_bp.route("/result-entry")
def result_entry():
    """Whole-Class Result Entry grid for bulk score entry"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    level_id = int_or_none(request.args.get("level_id"))
    class_id = int_or_none(request.args.get("class_id"))
    section_id = int_or_none(request.args.get("section_id"))
    
    selected_year = get_default_academic_year(year_id)
    selected_exam = db.session.get(Exam, exam_id) if exam_id else get_latest_exam_for_year(selected_year)
    if selected_exam and selected_year and selected_exam.academic_year_id != selected_year.id:
        selected_exam = get_latest_exam_for_year(selected_year)
    
    if not selected_year:
        flash("Please select an academic year.", "warning")
        return redirect(url_for("admin_advanced_results.new_dashboard"))

    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    exams = Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).all() if selected_year else []
    levels = AcademicLevel.query.filter_by(is_active=True).order_by(AcademicLevel.sort_order).all()
    
    # If no exam selected, show exam selection interface
    if not selected_exam:
        return render_template(
            "admin/result_entry.html",
            selected_year=selected_year,
            selected_exam=None,
            scope_info={},
            subjects=[],
            entry_grid=[],
            years=years,
            exams=exams,
            levels=levels,
            classes=[],
            sections=[],
            settings=get_settings(),
        )

    level_id = level_id or selected_exam.academic_level_id
    class_id = class_id or selected_exam.academic_class_id
    section_id = section_id or selected_exam.academic_section_id
    classes = AcademicClass.query.filter_by(academic_level_id=level_id, is_active=True).order_by(AcademicClass.sort_order, AcademicClass.name).all() if level_id else []
    sections = AcademicSection.query.filter_by(academic_class_id=class_id, is_active=True).order_by(AcademicSection.sort_order, AcademicSection.name).all() if class_id else []
    
    # Build scope info
    scope_info = {
        "level": db.session.get(AcademicLevel, level_id) if level_id else None,
        "class": db.session.get(AcademicClass, class_id) if class_id else None,
        "section": db.session.get(AcademicSection, section_id) if section_id else None,
    }
    
    subjects = subjects_for_scope(selected_exam, level_id=level_id, class_id=class_id)
    students = []
    if level_id and class_id:
        students = (
            students_for_scope_query(
                selected_year.id,
                level_id=level_id,
                class_id=class_id,
                section_id=section_id,
            )
            .order_by(Student.full_name)
            .all()
        )
    
    # Get existing results for these students and this exam
    student_ids = [s.id for s in students]
    existing_results = Result.query.filter(Result.student_id.in_(student_ids), Result.exam_id == selected_exam.id).all()
    results_dict = {(r.student_id, r.subject_id): r for r in existing_results}
    
    # Build entry grid data
    entry_grid = []
    for student in students:
        row_data = {
            "student": student,
            "results": {}
        }
        for subject in subjects:
            result = results_dict.get((student.id, subject.id))
            row_data["results"][subject.id] = {
                "result": result,
                "score": float(result.score) if result else None,
                "grade_override": result.grade_override if result else None,
                "is_published": result.is_published if result else True,
            }
        entry_grid.append(row_data)
    
    return render_template(
        "admin/result_entry.html",
        selected_year=selected_year,
        selected_exam=selected_exam,
        scope_info=scope_info,
        subjects=subjects,
        entry_grid=entry_grid,
        years=years,
        exams=exams,
        levels=levels,
        classes=classes,
        sections=sections,
        settings=get_settings(),
    )


@advanced_results_bp.route("/result-entry/autosave", methods=["POST"])
def autosave_result_entry():
    """Autosave a single score from the Results Hub entry grid."""
    year_id = int_or_none(request.form.get("year_id"))
    exam_id = int_or_none(request.form.get("exam_id"))
    level_id = int_or_none(request.form.get("level_id"))
    class_id = int_or_none(request.form.get("class_id"))
    section_id = int_or_none(request.form.get("section_id"))
    student_id = int_or_none(request.form.get("student_id"))
    subject_id = int_or_none(request.form.get("subject_id"))
    raw_score = request.form.get("score", "").strip()

    selected_year = db.session.get(AcademicYear, year_id)
    selected_exam = db.session.get(Exam, exam_id)
    student = db.session.get(Student, student_id)
    subject = db.session.get(Subject, subject_id)

    if not selected_year or not selected_exam or not student or not subject:
        return jsonify({"ok": False, "message": "Invalid result context."}), 400

    in_scope = students_for_scope_query(
        selected_year.id,
        level_id=level_id,
        class_id=class_id,
        section_id=section_id,
    ).filter(Student.id == student.id).first()
    if not in_scope:
        return jsonify({"ok": False, "message": "Student is outside the selected scope."}), 400

    if not raw_score:
        return jsonify({"ok": True, "status": "empty", "message": "No score entered."})

    try:
        score = float(raw_score)
    except ValueError:
        return jsonify({"ok": False, "message": "Invalid score."}), 400

    max_score = float(subject.max_score)
    if score < 0 or score > max_score:
        return jsonify({"ok": False, "message": f"Score must be between 0 and {max_score:g}."}), 400

    result = Result.query.filter_by(student_id=student.id, exam_id=selected_exam.id, subject_id=subject.id).first()
    if not result:
        result = Result(student=student, exam=selected_exam, subject=subject)
        db.session.add(result)

    result.score = score
    result.is_published = True
    audit("Result Entry", f"Autosaved {student.student_code} - {subject.name}: {score:g}")
    db.session.commit()

    return jsonify({
        "ok": True,
        "status": "saved",
        "score": score,
        "student_id": student.id,
        "subject_id": subject.id,
    })


@advanced_results_bp.route("/result-entry/save", methods=["POST"])
def save_result_entry():
    """Save bulk result entry from grid"""
    year_id = int_or_none(request.form.get("year_id"))
    exam_id = int_or_none(request.form.get("exam_id"))
    level_id = int_or_none(request.form.get("level_id"))
    class_id = int_or_none(request.form.get("class_id"))
    section_id = int_or_none(request.form.get("section_id"))
    
    selected_year = db.session.get(AcademicYear, year_id)
    selected_exam = db.session.get(Exam, exam_id)
    
    if not selected_year or not selected_exam:
        flash("Invalid selection.", "danger")
        return redirect(url_for("admin_advanced_results.new_dashboard"))
    
    subjects = subjects_for_scope(selected_exam, level_id=level_id, class_id=class_id)
    students = students_for_scope_query(
        selected_year.id,
        level_id=level_id,
        class_id=class_id,
        section_id=section_id,
    ).all()
    student_ids = [s.id for s in students]
    
    # Get existing results
    existing_results = Result.query.filter(Result.student_id.in_(student_ids), Result.exam_id == selected_exam.id).all()
    results_dict = {(r.student_id, r.subject_id): r for r in existing_results}
    
    # Process form data
    saved_count = 0
    validation_errors = []
    
    for student in students:
        for subject in subjects:
            score_key = f"score_{student.id}_{subject.id}"
            override_key = f"override_{student.id}_{subject.id}"
            published_key = f"published_{student.id}_{subject.id}"
            
            raw_score = request.form.get(score_key, "").strip()
            grade_override = request.form.get(override_key, "").strip()
            is_published = request.form.get(published_key) == "on"
            
            # Skip if no score entered
            if not raw_score:
                continue
            
            # Validate score against max_score
            try:
                score = float(raw_score)
                if score < 0 or score > float(subject.max_score):
                    validation_errors.append(f"{student.student_code} - {subject.name}: Score {score} exceeds max {subject.max_score}")
                    continue
            except ValueError:
                validation_errors.append(f"{student.student_code} - {subject.name}: Invalid score '{raw_score}'")
                continue
            
            # Get or create result
            result = results_dict.get((student.id, subject.id))
            if not result:
                result = Result(student=student, exam=selected_exam, subject=subject)
                db.session.add(result)
            
            result.score = score
            result.grade_override = grade_override if grade_override else None
            result.is_published = is_published
            saved_count += 1
    
    if validation_errors:
        flash(f"Saved {saved_count} results with {len(validation_errors)} validation errors.", "warning")
        for error in validation_errors[:5]:  # Show first 5 errors
            flash(error, "warning")
    else:
        flash(f"Successfully saved {saved_count} results.", "success")
    
    audit("Result Entry", f"Bulk saved {saved_count} results for exam {selected_exam.name}")
    db.session.commit()
    
    # Redirect back to entry grid with same scope
    return redirect(url_for("admin_advanced_results.result_entry", year_id=year_id, exam_id=exam_id, level_id=level_id, class_id=class_id, section_id=section_id))


@advanced_results_bp.route("/analytics")
def analytics():
    """Results analytics with charts matching design system"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    level_id = int_or_none(request.args.get("level_id"))
    class_id = int_or_none(request.args.get("class_id"))
    section_id = int_or_none(request.args.get("section_id"))
    subject_id = int_or_none(request.args.get("subject_id"))
    top_limit = int_or_none(request.args.get("top_limit")) or 5
    bottom_limit = int_or_none(request.args.get("bottom_limit")) or 5
    
    # Initialize selected variables before any conditional logic
    selected_year = None
    selected_exam = None
    selected_level = None
    selected_class = None
    selected_section = None
    selected_subject = None
    
    # Get selected year
    selected_year = db.session.get(AcademicYear, year_id) if year_id else AcademicYear.query.filter_by(is_current=True).first()
    
    if not selected_year:
        flash("Please select an academic year.", "warning")
        return redirect(url_for("admin_advanced_results.new_dashboard"))
    
    # Get selected exam from parameter or auto-select most recent
    selected_exam = db.session.get(Exam, exam_id) if exam_id else None
    if not selected_exam:
        selected_exam = Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).first()
    
    # Get filter data for selectors
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    exams = Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).all() if selected_year else []
    levels = AcademicLevel.query.filter_by(is_active=True).order_by(AcademicLevel.sort_order).all()
    classes = AcademicClass.query.order_by(AcademicClass.name).all()
    sections = AcademicSection.query.order_by(AcademicSection.name).all()
    subjects = Subject.query.order_by(Subject.name).all()
    
    # If still no exam available, show empty analytics
    if not selected_exam:
        # Provide empty analytics structure to prevent template errors
        empty_analytics = {
            "grade_distribution": {"labels": [], "counts": [], "colors": [], "total": 0},
            "subject_performance": {"labels": [], "scores": []},
            "exam_trend": {"labels": [], "scores": []},
            "pass_fail_ratio": {"pass": 0, "fail": 0, "total": 0},
            "overall_average": 0,
            "top_performers": [],
            "bottom_performers": [],
            "total_students": 0,
            "highest_score": 0,
            "lowest_score": 0,
        }
        return render_template(
            "admin/analytics.html",
            selected_year=selected_year,
            selected_exam=None,
            scope_info={},
            analytics=empty_analytics,
            years=years,
            exams=exams,
            levels=levels,
            classes=classes,
            sections=sections,
            subjects=subjects,
            selected_level_id=level_id,
            selected_class_id=class_id,
            selected_section_id=section_id,
            selected_subject_id=subject_id,
            top_limit=top_limit,
            bottom_limit=bottom_limit,
            settings=get_settings(),
        )
    
    # Build scope info
    scope_info = {
        "level": db.session.get(AcademicLevel, level_id) if level_id else None,
        "class": db.session.get(AcademicClass, class_id) if class_id else None,
        "section": db.session.get(AcademicSection, section_id) if section_id else None,
        "subject": db.session.get(Subject, subject_id) if subject_id else None,
    }
    
    students = (
        students_for_scope_query(
            selected_year.id,
            level_id=level_id,
            class_id=class_id,
            section_id=section_id,
        )
        .order_by(Student.full_name)
        .all()
    )
    student_ids = [s.id for s in students]
    
    # Get results for this exam
    results_query = Result.query.filter(Result.student_id.in_(student_ids), Result.exam_id == selected_exam.id, Result.is_published.is_(True))
    if scope_info["subject"]:
        results_query = results_query.filter_by(subject_id=scope_info["subject"].id)
    results = results_query.all()
    
    # Calculate analytics data with ranking limits
    analytics_data = build_analytics_data(results, students, selected_exam, top_limit, bottom_limit)
    
    return render_template(
        "admin/analytics.html",
        selected_year=selected_year,
        selected_exam=selected_exam,
        scope_info=scope_info,
        analytics=analytics_data,
        years=years,
        exams=exams,
        levels=levels,
        classes=classes,
        sections=sections,
        subjects=subjects,
        selected_level_id=level_id,
        selected_class_id=class_id,
        selected_section_id=section_id,
        selected_subject_id=subject_id,
        top_limit=top_limit,
        bottom_limit=bottom_limit,
        settings=get_settings(),
    )


def build_analytics_data(results, students, exam, top_limit=5, bottom_limit=5):
    """Build analytics data for charts using existing grade_for logic"""
    if not results:
        return {
            "grade_distribution": {"labels": [], "values": [], "colors": []},
            "subject_performance": {"labels": [], "values": []},
            "exam_trend": {"labels": [], "values": []},
            "pass_fail_ratio": {"pass": 0, "fail": 0, "total": 0},
            "overall_average": 0,
            "top_performers": [],
            "bottom_performers": [],
            "total_students": len(students) if students else 0,
            "highest_score": 0,
            "lowest_score": 0,
        }
    
    # Calculate percentages for each result
    percentages = []
    for result in results:
        if result.subject.max_score:
            pct = float(result.score) / float(result.subject.max_score) * 100
            percentages.append(pct)
    
    # Calculate overall average
    overall_average = round(sum(percentages) / len(percentages), 1) if percentages else 0
    
    # Calculate highest and lowest scores
    highest_score = round(max(percentages), 1) if percentages else 0
    lowest_score = round(min(percentages), 1) if percentages else 0
    
    # Grade distribution using actual grade scales
    grade_counts = {}
    grade_colors = {}
    for pct in percentages:
        grade_info = grade_for(pct, exam_id=exam.id)
        grade = grade_info["grade"]
        grade_counts[grade] = grade_counts.get(grade, 0) + 1
        if grade not in grade_colors:
            grade_colors[grade] = grade_info["badge_color"]
    
    grade_labels = sorted(grade_counts.keys())
    grade_values = [grade_counts[g] for g in grade_labels]
    grade_color_list = [grade_colors[g] for g in grade_labels]
    
    # Subject-wise performance
    subject_averages = {}
    for result in results:
        subject_name = result.subject.name
        pct = float(result.score) / float(result.subject.max_score) * 100 if result.subject.max_score else 0
        if subject_name not in subject_averages:
            subject_averages[subject_name] = []
        subject_averages[subject_name].append(pct)
    
    subject_labels = sorted(subject_averages.keys())
    subject_values = [round(sum(subject_averages[s]) / len(subject_averages[s]), 1) for s in subject_labels]
    
    # Exam trend (compare with other exams in same year)
    year_exams = Exam.query.filter_by(academic_year_id=exam.academic_year_id).order_by(Exam.created_at).all()
    exam_trend_labels = [e.name for e in year_exams]
    exam_trend_values = []
    for year_exam in year_exams:
        year_results = Result.query.filter(Result.exam_id == year_exam.id, Result.is_published.is_(True)).all()
        if year_results:
            year_pcts = [float(r.score) / float(r.subject.max_score) * 100 for r in year_results if r.subject.max_score]
            exam_trend_values.append(round(sum(year_pcts) / len(year_pcts), 1) if year_pcts else 0)
        else:
            exam_trend_values.append(0)
    
    # Pass/fail ratio
    pass_count = sum(1 for pct in percentages if grade_for(pct, exam_id=exam.id).get("is_pass"))
    fail_count = len(percentages) - pass_count
    
    # Student averages for top/bottom performers
    student_averages = {}
    for result in results:
        student_id = result.student_id
        pct = float(result.score) / float(result.subject.max_score) * 100 if result.subject.max_score else 0
        if student_id not in student_averages:
            student_averages[student_id] = []
        student_averages[student_id].append(pct)
    
    student_avg_list = [(sid, round(sum(pcts) / len(pcts), 1)) for sid, pcts in student_averages.items()]
    student_avg_list.sort(key=lambda x: x[1], reverse=True)
    
    top_performers = []
    bottom_performers = []
    
    for student_id, avg in student_avg_list[:top_limit]:
        student = next((s for s in students if s.id == student_id), None)
        if student:
            top_performers.append({
                "name": student.full_name,
                "mother_name": student.mother_name,
                "code": student.student_code,
                "average": avg,
                "grade": grade_for(avg, exam_id=exam.id)["grade"],
                "class_name": student.academic_class.name if student.academic_class else None,
                "section_name": student.academic_section.name if student.academic_section else None,
            })
    
    for student_id, avg in student_avg_list[-bottom_limit:]:
        student = next((s for s in students if s.id == student_id), None)
        if student:
            bottom_performers.append({
                "name": student.full_name,
                "mother_name": student.mother_name,
                "code": student.student_code,
                "average": avg,
                "grade": grade_for(avg, exam_id=exam.id)["grade"],
                "class_name": student.academic_class.name if student.academic_class else None,
                "section_name": student.academic_section.name if student.academic_section else None,
            })
    
    return {
        "grade_distribution": {
            "labels": grade_labels,
            "counts": grade_values,
            "colors": grade_color_list,
            "total": sum(grade_values) if grade_values else 0,
        },
        "subject_performance": {
            "labels": subject_labels,
            "scores": subject_values,
        },
        "exam_trend": {
            "labels": exam_trend_labels,
            "scores": exam_trend_values,
        },
        "pass_fail_ratio": {
            "pass": pass_count,
            "fail": fail_count,
            "total": len(percentages),
        },
        "overall_average": overall_average,
        "top_performers": top_performers,
        "bottom_performers": bottom_performers,
        "total_students": len(students) if students else 0,
        "highest_score": highest_score,
        "lowest_score": lowest_score,
    }


@advanced_results_bp.route("/grade-management")
def grade_management():
    """Grade Management page scoped per exam"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    
    # Get selected year and exam
    selected_year = db.session.get(AcademicYear, year_id) if year_id else AcademicYear.query.filter_by(is_current=True).first()
    selected_exam = db.session.get(Exam, exam_id) if exam_id else None
    
    # Get all years and exams for selectors
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    exams = Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).all() if selected_year else []
    
    # Get grade scales for the selected exam (or global if no exam selected)
    if selected_exam:
        grade_scales = GradeScale.query.filter_by(exam_id=exam_id).order_by(GradeScale.sort_order.asc(), GradeScale.min_score.desc()).all()
        using_global = len(grade_scales) == 0
        if using_global:
            grade_scales = GradeScale.query.filter_by(exam_id=None).order_by(GradeScale.sort_order.asc(), GradeScale.min_score.desc()).all()
    else:
        grade_scales = GradeScale.query.filter_by(exam_id=None).order_by(GradeScale.sort_order.asc(), GradeScale.min_score.desc()).all()
        using_global = True
    
    # Get all exams with their configuration status
    all_exams = Exam.query.order_by(Exam.id.desc()).all()
    exam_status = []
    for exam in all_exams:
        has_custom_scales = GradeScale.query.filter_by(exam_id=exam.id).count() > 0
        exam_status.append({
            "exam": exam,
            "configured": has_custom_scales,
        })
    
    # Calculate total points for selected exam
    total_points = 0
    if selected_exam and selected_exam.academic_level_id:
        subjects = Subject.query.filter_by(academic_level_id=selected_exam.academic_level_id).all()
        total_points = sum(float(s.max_score) for s in subjects)
    
    return render_template(
        "admin/grade_management.html",
        years=years,
        exams=exams,
        selected_year=selected_year,
        selected_exam=selected_exam,
        grade_scales=grade_scales,
        using_global=using_global,
        exam_status=exam_status,
        total_points=total_points,
        settings=get_settings(),
    )


@advanced_results_bp.route("/grade-management/save", methods=["POST"])
def save_grade_scales():
    """Save grade scales for an exam"""
    exam_id = int_or_none(request.form.get("exam_id"))
    
    # Update existing grade scales
    for grade in GradeScale.query.filter_by(exam_id=exam_id if exam_id else None).all():
        grade.grade = request.form.get(f"grade_{grade.id}", grade.grade).strip() or grade.grade
        grade.min_score = request.form.get(f"min_{grade.id}", grade.min_score)
        grade.max_score = request.form.get(f"max_{grade.id}", grade.max_score)
        grade.grade_point = request.form.get(f"point_{grade.id}", grade.grade_point)
        grade.comment = request.form.get(f"comment_{grade.id}", grade.comment).strip() or grade.comment
        grade.is_pass = request.form.get(f"status_{grade.id}", "fail") == "pass"
        grade.badge_color = request.form.get(f"badge_color_{grade.id}", grade.badge_color)
        grade.text_color = request.form.get(f"text_color_{grade.id}", grade.text_color)
        grade.background_color = request.form.get(f"background_color_{grade.id}", grade.background_color)
        grade.border_color = request.form.get(f"border_color_{grade.id}", grade.border_color)
        grade.sort_order = int(request.form.get(f"sort_order_{grade.id}") or grade.sort_order or 0)
        grade.is_active = request.form.get(f"active_{grade.id}") == "on"
    
    # Create new grade scale if provided
    new_grade = request.form.get("new_grade", "").strip()
    if new_grade and exam_id:
        db.session.add(GradeScale(
            grade=new_grade,
            exam_id=exam_id,
            min_score=request.form.get("new_min") or 0,
            max_score=request.form.get("new_max") or 0,
            grade_point=request.form.get("new_point") or 0,
            comment=request.form.get("new_comment", "").strip() or "Custom grade",
            is_pass=request.form.get("new_status", "pass") == "pass",
            badge_color=request.form.get("new_badge_color") or "#2563eb",
            text_color=request.form.get("new_text_color") or "#ffffff",
            background_color=request.form.get("new_background_color") or "#eff6ff",
            border_color=request.form.get("new_border_color") or "#3b82f6",
            sort_order=0,
            is_active=True,
        ))
    
    audit("Grade Management", f"Saved grade scales for exam {exam_id if exam_id else 'global'}")
    db.session.commit()
    flash("Grade scales saved successfully.", "success")
    
    if exam_id:
        return redirect(url_for("admin_advanced_results.grade_management", exam_id=exam_id))
    return redirect(url_for("admin_advanced_results.grade_management"))


@advanced_results_bp.route("/settings")
def results_settings():
    """Results Settings page with label and language customization"""
    settings = get_settings()
    
    # Get all unique label keys
    label_keys = db.session.query(LabelTranslation.label_key).distinct().order_by(LabelTranslation.label_key).all()
    label_keys = [k[0] for k in label_keys]
    
    # Get all unique language codes
    language_codes = db.session.query(LabelTranslation.language_code).distinct().order_by(LabelTranslation.language_code).all()
    language_codes = [k[0] for k in language_codes]
    
    # Build label translations matrix
    label_matrix = {}
    for key in label_keys:
        label_matrix[key] = {}
        for lang in language_codes:
            translation = LabelTranslation.query.filter_by(label_key=key, language_code=lang).first()
            label_matrix[key][lang] = {
                'text_value': translation.text_value if translation else '',
                'context': translation.context if translation else '',
                'id': translation.id if translation else None
            }
    
    return render_template(
        "admin/results_settings.html",
        label_keys=label_keys,
        language_codes=language_codes,
        label_matrix=label_matrix,
        settings=settings,
    )


@advanced_results_bp.route("/students-management")
def students_management():
    """Students Management page within Results Hub - consolidated student operations"""
    year_id = int_or_none(request.args.get("year_id"))
    level_id = int_or_none(request.args.get("level_id"))
    class_id = int_or_none(request.args.get("class_id"))
    search_query = request.args.get("q", "").strip()
    status_filter = request.args.get("status_filter", "")
    page = int_or_none(request.args.get("page", 1)) or 1
    per_page = 7  # Students per page for balanced layout
    
    # Get selected year
    selected_year = db.session.get(AcademicYear, year_id) if year_id else AcademicYear.query.filter_by(is_current=True).first()
    
    # Get all years for selector
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    
    # Get levels and classes for filters
    levels = AcademicLevel.query.filter_by(is_active=True).order_by(AcademicLevel.sort_order).all()
    classes = AcademicClass.query.order_by(AcademicClass.name).all() if not level_id else AcademicClass.query.filter_by(academic_level_id=level_id).all()
    
    # Calculate statistics for summary cards
    stats = {
        "total_students": 0,
        "secondary": 0,
        "upper_primary": 0,
        "lower_primary": 0,
        "kindergarten": 0,
        "active_students": 0,
    }
    
    if selected_year:
        # Total students for selected year
        stats["total_students"] = Student.query.filter_by(academic_year_id=selected_year.id).count()
        
        # Students by level
        for level in levels:
            level_count = Student.query.filter_by(academic_year_id=selected_year.id, academic_level_id=level.id).count()
            level_name_lower = level.name.lower()
            if "kindergarten" in level_name_lower or "kg" in level_name_lower:
                stats["kindergarten"] = level_count
            elif "upper" in level_name_lower and "primary" in level_name_lower:
                stats["upper_primary"] = level_count
            elif "lower" in level_name_lower and "primary" in level_name_lower:
                stats["lower_primary"] = level_count
            elif "primary" in level_name_lower:
                # If just "primary" without upper/lower, count as upper primary
                stats["upper_primary"] += level_count
            elif "middle" in level_name_lower:
                stats["upper_primary"] += level_count  # Count middle as upper primary for this context
            elif "secondary" in level_name_lower:
                stats["secondary"] = level_count
        
        # Active students (not locked)
        stats["active_students"] = Student.query.filter_by(academic_year_id=selected_year.id, is_result_locked=False).count()
    
    # Get students with filters
    students_query = Student.query.filter_by(academic_year_id=selected_year.id) if selected_year else Student.query()
    
    if level_id:
        students_query = students_query.filter_by(academic_level_id=level_id)
    if class_id:
        students_query = students_query.filter_by(academic_class_id=class_id)
    
    # Apply search filter
    if search_query:
        search_pattern = f"%{search_query}%"
        students_query = students_query.filter(
            db.or_(
                Student.student_code.like(search_pattern),
                Student.full_name.like(search_pattern),
                Student.mother_name.like(search_pattern)
            )
        )
    
    # Apply status filter
    if status_filter == "locked":
        students_query = students_query.filter_by(is_result_locked=True)
    elif status_filter == "active":
        students_query = students_query.filter_by(is_result_locked=False)
    
    # Get total count for pagination
    total_students = students_query.count()
    
    # Apply pagination
    students = students_query.order_by(Student.student_code).offset((page - 1) * per_page).limit(per_page).all()
    
    # Calculate pagination info
    total_pages = (total_students + per_page - 1) // per_page if total_students > 0 else 1
    has_prev = page > 1
    has_next = page < total_pages
    
    return render_template(
        "admin/students_management.html",
        years=years,
        selected_year=selected_year,
        levels=levels,
        classes=classes,
        students=students,
        selected_level_id=level_id,
        selected_class_id=class_id,
        q=search_query,
        status_filter=status_filter,
        stats=stats,
        page=page,
        per_page=per_page,
        total_students=total_students,
        total_pages=total_pages,
        has_prev=has_prev,
        has_next=has_next,
        settings=get_settings(),
    )


@advanced_results_bp.route("/students/new", methods=["GET", "POST"])
@advanced_results_bp.route("/students/<int:student_id>/edit", methods=["GET", "POST"])
def student_form(student_id=None):
    """Unified Results Hub student create/edit form."""
    student = db.session.get(Student, student_id) if student_id else Student()
    if request.method == "POST":
        try:
            save_student_from_form(student)
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(request.url)
        db.session.add(student)
        audit("Student Updates", f"Saved student {student.student_code}")
        db.session.commit()
        flash("Student saved successfully.", "success")
        return redirect(url_for("admin_advanced_results.students_management", year_id=student.academic_year_id, level_id=student.academic_level_id, class_id=student.academic_class_id))

    academic_levels = AcademicLevel.query.filter_by(is_active=True).order_by(AcademicLevel.sort_order).all()
    incident_reports = IncidentReport.query.filter_by(student_id=student.id).order_by(IncidentReport.created_at.desc()).limit(10).all() if student.id else []
    return render_template(
        "admin/student_form.html",
        student=student,
        years=AcademicYear.query.order_by(AcademicYear.name.desc()).all(),
        academic_levels=academic_levels,
        incident_reports=incident_reports,
        student_form_action=url_for("admin_advanced_results.student_form", student_id=student.id) if student.id else url_for("admin_advanced_results.student_form"),
    )


@advanced_results_bp.route("/students/<int:student_id>/delete", methods=["POST"])
def delete_student(student_id):
    student = db.session.get(Student, student_id) or abort(404)
    db.session.delete(student)
    audit("Student Updates", f"Deleted student {student.student_code}")
    db.session.commit()
    flash("Student deleted.", "success")
    return redirect(url_for("admin_advanced_results.students_management"))


@advanced_results_bp.route("/students/<int:student_id>/data")
def student_data_json(student_id):
    """API endpoint to fetch student data as JSON for AJAX preview"""
    student = db.session.get(Student, student_id) or abort(404)
    
    # Construct photo URL from photo_path if it exists
    photo_url = None
    if student.photo_path:
        # If photo_path already starts with /static/, use it as-is
        if student.photo_path.startswith('/static/'):
            photo_url = student.photo_path
        # If photo_path is relative to uploads folder, prepend /static/
        elif student.photo_path.startswith('uploads/'):
            photo_url = f"/static/{student.photo_path}"
        # Otherwise, assume it's a relative path and prepend /static/uploads/
        else:
            photo_url = f"/static/uploads/{student.photo_path}"
    
    return jsonify({
        "id": student.id,
        "student_code": student.student_code,
        "full_name": student.full_name,
        "mother_name": student.mother_name,
        "phone": student.phone,
        "photo_path": student.photo_path,
        "photo_url": photo_url,
        "academic_level_id": student.academic_level_id,
        "academic_level_name": student.academic_level.name if student.academic_level else None,
        "academic_class_id": student.academic_class_id,
        "academic_class_name": student.academic_class.name if student.academic_class else None,
        "academic_section_id": student.academic_section_id,
        "academic_section_name": student.academic_section.name if student.academic_section else None,
        "academic_year_id": student.academic_year_id,
        "academic_year_name": student.academic_year.name if student.academic_year else None,
        "is_result_locked": student.is_result_locked,
        "is_active": student.is_active,
    })


@advanced_results_bp.route("/students/<int:student_id>/toggle-lock", methods=["POST"])
def toggle_student_lock(student_id):
    student = db.session.get(Student, student_id) or abort(404)
    if student.is_result_locked and not can("unlock_results"):
        abort(403)
    if not student.is_result_locked and not can("lock_results"):
        abort(403)
    student.is_result_locked = not student.is_result_locked
    if student.is_result_locked:
        student.lock_reason = request.form.get("lock_reason", "").strip() or "Outstanding clearance required."
        audit("Result Locking", f"Locked result for {student.student_code}")
    else:
        student.lock_reason = ""
        audit("Result Locking", f"Unlocked result for {student.student_code}")
    db.session.commit()
    flash("Result lock status updated.", "success")
    return redirect(url_for("admin_advanced_results.students_management", year_id=student.academic_year_id, level_id=student.academic_level_id, class_id=student.academic_class_id))


@advanced_results_bp.route("/students/import", methods=["POST"])
def import_students():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename, ALLOWED_SHEETS):
        flash("Upload an .xlsx file.", "danger")
        return redirect(url_for("admin_advanced_results.students_management"))
    rows, errors = preview_students(file)
    if not errors:
        session["student_import_rows"] = rows
    audit("Import Operations", f"Previewed student import: {len(rows)} rows, {len(errors)} errors")
    db.session.commit()
    return render_template("admin/import_wizard.html", kind="students", rows=rows, errors=errors, confirm_url=url_for("admin_advanced_results.confirm_student_import"))


@advanced_results_bp.route("/students/import/confirm", methods=["POST"])
def confirm_student_import():
    rows = session.pop("student_import_rows", [])
    if not rows:
        flash("No validated student import is waiting for confirmation.", "warning")
        return redirect(url_for("admin_advanced_results.students_management"))
    try:
        for data in rows:
            year = AcademicYear.query.filter_by(name=data["academic_year"]).one()
            student = Student.query.filter_by(student_code=data["student_id"]).first() or Student(student_code=data["student_id"])
            student.full_name = data["full_name"]
            student.mother_name = data.get("mother_name", "")
            student.phone = data.get("phone", "")
            student.academic_year = year
            map_imported_student_class(student, data["class"])
            student.is_active = True
            db.session.add(student)
        audit("Import Operations", f"Confirmed student import: {len(rows)} rows")
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash("Import failed. No records were saved.", "danger")
        return redirect(url_for("admin_advanced_results.students_management"))
    flash(f"Imported {len(rows)} students.", "success")
    return redirect(url_for("admin_advanced_results.students_management"))


@advanced_results_bp.route("/students/import/template")
def student_import_template():
    return workbook_response(student_template(), "student_import_template.xlsx")


@advanced_results_bp.route("/students/export")
def export_students():
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["student_id", "full_name", "mother_name", "phone", "class", "academic_year", "active"])
    for student in Student.query.order_by(Student.full_name).all():
        class_name = student.academic_class.name if student.academic_class else (student.school_class.name if student.school_class else "")
        year_name = student.academic_year.name if student.academic_year else ""
        ws.append([student.student_code, student.full_name, student.mother_name, student.phone, class_name, year_name, student.is_active])
    return workbook_response(wb, "students.xlsx")


def save_student_from_form(student):
    student.student_code = request.form["student_code"].strip()
    student.full_name = request.form["full_name"].strip()
    student.mother_name = request.form.get("mother_name", "").strip()
    student.phone = request.form.get("phone", "").strip()
    student.academic_year_id = int_or_none(request.form.get("academic_year_id"))
    student.academic_level_id = int_or_none(request.form.get("academic_level_id"))
    student.academic_class_id = int_or_none(request.form.get("academic_class_id"))
    student.academic_section_id = int_or_none(request.form.get("academic_section_id"))
    if not student.academic_class_id and not student.class_id:
        raise ValueError("Please select a class for this student.")
    sync_student_legacy_class(student)
    student.note = request.form.get("note", "").strip()
    student.is_result_locked = bool(request.form.get("is_result_locked"))
    student.lock_reason = request.form.get("lock_reason", "").strip()
    student.is_active = bool(request.form.get("is_active"))
    photo = request.files.get("photo")
    if photo and photo.filename:
        if not allowed_file(photo.filename, ALLOWED_PHOTOS):
            raise ValueError("Photo must be JPG, PNG, or WEBP.")
        student.photo_path = upload_image(photo, "school/students")


def sync_student_legacy_class(student):
    if not student.academic_class_id:
        return
    academic_class = db.session.get(AcademicClass, student.academic_class_id)
    if not academic_class:
        return
    school_class = SchoolClass.query.filter_by(name=academic_class.name).first()
    if not school_class:
        school_class = SchoolClass(name=academic_class.name)
        db.session.add(school_class)
        db.session.flush()
    student.school_class = school_class
    student.level = academic_class.academic_level.name if academic_class.academic_level else student.level
    if student.academic_section_id:
        section = db.session.get(AcademicSection, student.academic_section_id)
        student.section = section.name if section else student.section


def map_imported_student_class(student, class_name):
    school_class = SchoolClass.query.filter_by(name=class_name).first()
    if not school_class:
        school_class = SchoolClass(name=class_name)
        db.session.add(school_class)
        db.session.flush()
    student.school_class = school_class

    academic_class = AcademicClass.query.filter_by(name=class_name).first()
    if academic_class:
        student.academic_class = academic_class
        student.academic_level = academic_class.academic_level
        student.level = academic_class.academic_level.name if academic_class.academic_level else student.level


@advanced_results_bp.route("/settings/save-labels", methods=["POST"])
def save_label_translations():
    """Save label translations"""
    default_language = request.form.get("default_language", "").strip()
    if default_language:
        setting = db.session.get(Setting, "default_language") or Setting(key="default_language")
        setting.value = default_language
        db.session.add(setting)

    new_language = request.form.get("new_language", "").strip().lower()
    if new_language:
        for label_key, _, text_value, context in RESULTS_LABEL_SEEDS:
            exists = LabelTranslation.query.filter_by(label_key=label_key, language_code=new_language).first()
            if not exists:
                db.session.add(LabelTranslation(
                    label_key=label_key,
                    language_code=new_language,
                    text_value=text_value,
                    context=context,
                ))

    for key, value in request.form.items():
        if key.startswith('label_'):
            parts = key.split('_')
            if len(parts) >= 3:
                label_key = '_'.join(parts[1:-1])
                language_code = parts[-1]
                text_value = value.strip()
                
                # Find existing translation or create new
                translation = LabelTranslation.query.filter_by(label_key=label_key, language_code=language_code).first()
                if translation:
                    translation.text_value = text_value
                else:
                    db.session.add(LabelTranslation(
                        label_key=label_key,
                        language_code=language_code,
                        text_value=text_value,
                        context='Results Settings'
                    ))
    
    db.session.commit()
    flash("Label translations saved successfully.", "success")
    return redirect(url_for("admin_advanced_results.results_settings"))


def result_filters():
    return {
        "q": request.args.get("q", "").strip(),
        "year_id": int_or_none(request.args.get("year_id")),
        "exam_id": int_or_none(request.args.get("exam_id")),
        "class_id": int_or_none(request.args.get("class_id")),
        "level": request.args.get("level", "").strip(),
        "section": request.args.get("section", "").strip(),
        "status": request.args.get("status", "").strip(),
    }


def result_query(filters):
    query = Result.query.join(Result.student).join(Result.exam).join(Result.subject)
    if filters["q"]:
        q = f"%{filters['q']}%"
        query = query.filter(or_(Student.student_code.like(q), Student.full_name.like(q), Subject.name.like(q)))
    if filters["year_id"]:
        query = query.filter(Exam.academic_year_id == filters["year_id"])
    if filters["exam_id"]:
        query = query.filter(Result.exam_id == filters["exam_id"])
    if filters["class_id"]:
        query = query.filter(Student.class_id == filters["class_id"])
    if filters["level"]:
        query = query.filter(Student.level == filters["level"])
    if filters["section"]:
        query = query.filter(Student.section == filters["section"])
    if filters["status"] == "Published":
        query = query.filter(Result.is_published.is_(True))
    elif filters["status"] == "Locked":
        query = query.filter(Student.is_result_locked.is_(True))
    elif filters["status"] == "Unpublished":
        query = query.filter(Result.is_published.is_(False))
    return query


def group_results(rows):
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list)))))
    for row in rows:
        grouped[row.exam.academic_year.name][row.exam.name][row.student.level or "No Level"][row.student.school_class.name][row.student.section or "No Section"].append(row)
    return grouped


def build_stats(payloads, rows):
    averages = [Decimal(str(payload["average"])) for payload in payloads if payload.get("subjects")]
    pass_count = sum(1 for avg in averages if grade_for(avg).get("is_pass"))
    subject_totals = defaultdict(list)
    for row in rows:
        subject_totals[row.subject.name].append(Decimal(row.score) / Decimal(row.subject.max_score) * 100 if row.subject.max_score else Decimal("0"))
    subject_averages = {name: round(sum(values) / len(values), 2) for name, values in subject_totals.items() if values}
    ranked = sorted(payloads, key=lambda item: item.get("average", 0), reverse=True)
    return {
        "rows": len(rows),
        "students": len({row.student_id for row in rows}),
        "published": sum(1 for row in rows if row.is_published),
        "locked": len({row.student_id for row in rows if row.student.is_result_locked}),
        "pass_rate": round(pass_count / len(averages) * 100, 2) if averages else 0,
        "fail_rate": round((len(averages) - pass_count) / len(averages) * 100, 2) if averages else 0,
        "top_students": ranked[:5],
        "lowest_students": list(reversed(ranked[-5:])),
        "subject_averages": subject_averages,
        "exam_average": round(sum(averages) / len(averages), 2) if averages else 0,
    }


def distinct_values(column):
    return [value[0] for value in db.session.query(column).filter(column.isnot(None), column != "").distinct().order_by(column).all()]


def int_or_none(value):
    return int(value) if value and str(value).isdigit() else None


def workbook_response(workbook, filename):
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    workbook.save(tmp.name)
    tmp.close()
    return send_file(
        tmp.name,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def build_dashboard_stats(exam):
    """Build statistics for the dashboard"""
    students_query = students_for_scope_query(exam.academic_year_id, exam=exam)
    student_ids = [student.id for student in students_query.all()]
    total_students = len(student_ids)

    subjects = subjects_for_scope(exam)
    total_subjects = len(subjects)
    
    expected_results = total_students * total_subjects if total_subjects > 0 else 0
    actual_results = 0
    if student_ids:
        actual_results = Result.query.filter(
            Result.exam_id == exam.id,
            Result.student_id.in_(student_ids),
            Result.is_published.is_(True),
        ).count()
    completion_percentage = round((actual_results / expected_results * 100), 2) if expected_results > 0 else 0
    
    # Get active classes
    if exam.academic_class_id:
        active_classes = 1
    elif exam.academic_level_id:
        active_classes = AcademicClass.query.filter_by(academic_level_id=exam.academic_level_id, is_active=True).count()
    else:
        active_classes = AcademicClass.query.filter_by(is_active=True).count()
    
    return {
        "total_students": total_students,
        "total_subjects": total_subjects,
        "completion_percentage": completion_percentage,
        "active_classes": active_classes,
        "expected_results": expected_results,
        "actual_results": actual_results,
    }


def build_class_cards(exam, level_filter=None):
    """Build class cards for the dashboard, optionally filtered by level"""
    cards = []
    
    # Determine scope based on exam configuration
    if exam.academic_section_id:
        # Single section - show section card
        section = db.session.get(AcademicSection, exam.academic_section_id)
        if section:
            cards.append(build_single_class_card(exam, section=section))
    elif exam.academic_class_id:
        # Single class - show section cards within it
        academic_class = db.session.get(AcademicClass, exam.academic_class_id)
        if academic_class:
            sections = AcademicSection.query.filter_by(academic_class_id=academic_class.id, is_active=True).all()
            for section in sections:
                cards.append(build_single_class_card(exam, section=section))
    elif exam.academic_level_id:
        # Single level - show class cards
        classes = AcademicClass.query.filter_by(academic_level_id=exam.academic_level_id, is_active=True).all()
        for cls in classes:
            cards.append(build_single_class_card(exam, academic_class=cls))
    else:
        # No scope - show level cards
        levels = AcademicLevel.query.filter_by(is_active=True).order_by(AcademicLevel.sort_order).all()
        for level in levels:
            cards.append(build_single_class_card(exam, academic_level=level))
    
    # Apply level filter if provided
    if level_filter:
        cards = [card for card in cards if card.get('academic_level_id') == level_filter.id]
    
    return cards


def build_single_class_card(exam, academic_level=None, academic_class=None, section=None):
    """Build a single class/level/section card"""
    # Determine label based on what's provided
    if section:
        if not academic_class:
            academic_class = section.academic_class
        label = f"{academic_class.name} - {section.name}" if academic_class else section.name
        student_query = students_for_scope_query(
            academic_year_id=exam.academic_year_id,
            class_id=academic_class.id if academic_class else None,
            section_id=section.id,
        )
    elif academic_class:
        label = academic_class.name
        student_query = students_for_scope_query(
            academic_year_id=exam.academic_year_id,
            class_id=academic_class.id,
        )
    elif academic_level:
        label = academic_level.name
        student_query = students_for_scope_query(
            academic_year_id=exam.academic_year_id,
            level_id=academic_level.id,
        )
    else:
        label = "All Students"
        student_query = students_for_scope_query(exam.academic_year_id)
    
    student_count = student_query.count()
    
    # Calculate completion for this scope (only published results)
    student_ids = [s.id for s in student_query.all()]
    if student_ids:
        subjects = subjects_for_scope(
            exam,
            level_id=academic_level.id if academic_level else None,
            class_id=academic_class.id if academic_class else None,
        )
        expected = len(student_ids) * len(subjects)
        actual = Result.query.filter(Result.exam_id == exam.id, Result.student_id.in_(student_ids), Result.is_published.is_(True)).count()
        completion = round((actual / expected * 100), 2) if expected > 0 else 0
    else:
        completion = 0
    
    return {
        "label": label,
        "student_count": student_count,
        "completion_percentage": completion,
        "academic_level_id": academic_level.id if academic_level else None,
        "academic_class_id": academic_class.id if academic_class else None,
        "academic_section_id": section.id if section else None,
    }
